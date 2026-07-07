"""
Export engine for the dashboard's Export buttons.
Three formats, each laid out for its own medium rather than one dump reused three ways:
  - CSV   : flat, for spreadsheets/scripts to consume further
  - XLSX  : multi-sheet workbook, styled headers, frozen panes — for opening and reading directly
  - PDF   : a branded one-page executive report — for sharing/printing
"""
import io
import json
import sqlite3
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

ROOT = Path(__file__).resolve().parent.parent
SRC_DIR = ROOT / "src"
import sys  # noqa: E402
sys.path.insert(0, str(SRC_DIR))
from tenant import tenant_paths  # noqa: E402

# brand colors, matched to frontend/assets/css/theme.css
INK = colors.HexColor("#0b0d10")
GOLD = colors.HexColor("#c9a227")
TEAL = colors.HexColor("#2c7a73")
MUTED = colors.HexColor("#5b5f66")
LIGHT_ROW = colors.HexColor("#f2efe7")


def _load_data(client_id: str | None = None):
    paths = tenant_paths(client_id)
    conn = sqlite3.connect(paths["db"])
    summary = pd.read_sql("SELECT * FROM daily_summary ORDER BY day ASC", conn)
    transactions = pd.read_sql("SELECT * FROM transactions ORDER BY tx_time DESC", conn)
    conn.close()

    analytics = json.loads(paths["analytics"].read_text()) if paths["analytics"].exists() else {}
    quality = json.loads(paths["quality"].read_text()) if paths["quality"].exists() else {}
    return summary, transactions, analytics, quality


# ---------------------------------------------------------------- CSV
def build_csv(client_id: str | None = None) -> bytes:
    """Flat CSV: transaction-level data, the most re-usable shape for further analysis."""
    _, transactions, _, _ = _load_data(client_id)
    buf = io.StringIO()
    transactions.to_csv(buf, index=False)
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------- XLSX
def build_xlsx(client_id: str | None = None) -> bytes:
    summary, transactions, analytics, quality = _load_data(client_id)

    wb = Workbook()

    header_fill = PatternFill(start_color="0B0D10", end_color="0B0D10", fill_type="solid")
    header_font = Font(color="C9A227", bold=True, name="Calibri", size=11)
    title_font = Font(bold=True, size=14, color="0B0D10")

    def style_header(ws, row=1, ncols=1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def autosize(ws, ncols):
        for c in range(1, ncols + 1):
            letter = get_column_letter(c)
            max_len = max((len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, ws.max_row + 1)), default=10)
            ws.column_dimensions[letter].width = min(max(max_len + 3, 10), 40)

    # --- Sheet 1: Summary ---
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Blackloom — Pipeline Summary"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}"
    ws1["A2"].font = Font(italic=True, size=9, color="5B5F66")

    ws1["A4"] = "Overall quality score"
    ws1["B4"] = quality.get("overall_score", "N/A")
    ws1["A5"] = "Total daily-summary rows"
    ws1["B5"] = len(summary)
    ws1["A6"] = "Total transactions"
    ws1["B6"] = len(transactions)

    start_row = 8
    ws1.cell(row=start_row, column=1, value="Day")
    ws1.cell(row=start_row, column=2, value="Tx Count")
    ws1.cell(row=start_row, column=3, value="Total ETH Volume")
    ws1.cell(row=start_row, column=4, value="Avg Gas (gwei)")
    ws1.cell(row=start_row, column=5, value="Failed Tx")
    style_header(ws1, row=start_row, ncols=5)
    for i, row in summary.iterrows():
        r = start_row + 1 + i
        ws1.cell(row=r, column=1, value=row["day"])
        ws1.cell(row=r, column=2, value=int(row["tx_count"]))
        ws1.cell(row=r, column=3, value=round(float(row["total_eth_volume"]), 4))
        ws1.cell(row=r, column=4, value=round(float(row["avg_gas_price_gwei"]), 2))
        ws1.cell(row=r, column=5, value=int(row["failed_tx_count"]))
    ws1.freeze_panes = f"A{start_row + 1}"
    autosize(ws1, 5)

    # --- Sheet 2: Transactions ---
    ws2 = wb.create_sheet("Transactions")
    cols = ["hash", "from_address", "to_address", "value_eth", "gas_price_gwei", "tx_time", "block_number", "is_error"]
    for c, name in enumerate(cols, start=1):
        ws2.cell(row=1, column=c, value=name)
    style_header(ws2, row=1, ncols=len(cols))
    for i, row in transactions.iterrows():
        for c, name in enumerate(cols, start=1):
            ws2.cell(row=i + 2, column=c, value=row[name])
    ws2.freeze_panes = "A2"
    autosize(ws2, len(cols))

    # --- Sheet 3: Analytics ---
    ws3 = wb.create_sheet("Analytics")
    ws3["A1"] = "Anomalies (z-score > 3)"
    ws3["A1"].font = Font(bold=True)
    anomalies = analytics.get("anomalies", [])
    headers = ["hash", "metric", "value", "z_score", "tx_time"]
    for c, h in enumerate(headers, start=1):
        ws3.cell(row=2, column=c, value=h)
    style_header(ws3, row=2, ncols=len(headers))
    for i, a in enumerate(anomalies):
        for c, h in enumerate(headers, start=1):
            ws3.cell(row=3 + i, column=c, value=a.get(h))

    forecast_start = 3 + len(anomalies) + 2
    ws3.cell(row=forecast_start, column=1, value="3-Day Forecast").font = Font(bold=True)
    ws3.cell(row=forecast_start + 1, column=1, value="Day")
    ws3.cell(row=forecast_start + 1, column=2, value="Predicted Volume (ETH)")
    style_header(ws3, row=forecast_start + 1, ncols=2)
    for i, f in enumerate(analytics.get("forecast", [])):
        ws3.cell(row=forecast_start + 2 + i, column=1, value=f["day"])
        ws3.cell(row=forecast_start + 2 + i, column=2, value=f["predicted_volume"])
    autosize(ws3, 5)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- PDF
def build_pdf(client_id: str | None = None) -> bytes:
    summary, transactions, analytics, quality = _load_data(client_id)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=22 * mm, bottomMargin=18 * mm,
                             leftMargin=18 * mm, rightMargin=18 * mm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("BLTitle", parent=styles["Title"], fontName="Helvetica-Bold",
                                  fontSize=22, textColor=INK, spaceAfter=2)
    eyebrow_style = ParagraphStyle("BLEyebrow", parent=styles["Normal"], fontName="Helvetica",
                                    fontSize=9, textColor=TEAL, spaceAfter=14, letterSpacing=1.2)
    h2_style = ParagraphStyle("BLH2", parent=styles["Heading2"], fontName="Helvetica-Bold",
                               fontSize=13, textColor=INK, spaceBefore=16, spaceAfter=8)
    body_style = ParagraphStyle("BLBody", parent=styles["Normal"], fontSize=9.5, textColor=colors.black, leading=13)
    muted_style = ParagraphStyle("BLMuted", parent=styles["Normal"], fontSize=8, textColor=MUTED)

    elements = [
        Paragraph("BLACKLOOM", title_style),
        Paragraph("ON-CHAIN ETL &amp; ANALYTICS — PIPELINE REPORT", eyebrow_style),
        Paragraph(f"Generated {datetime.now().strftime('%B %d, %Y at %H:%M UTC')}", muted_style),
        Spacer(1, 14),
    ]

    # KPI row
    total_tx = len(transactions)
    total_vol = round(transactions["value_eth"].sum(), 4) if len(transactions) else 0
    failed = int(transactions["is_error"].sum()) if len(transactions) else 0
    overall_q = quality.get("overall_score", "N/A")

    kpi_data = [["Total Transactions", "Total ETH Volume", "Failed Tx", "Quality Score"],
                [f"{total_tx:,}", f"{total_vol:,}", str(failed), f"{overall_q}%"]]
    kpi_table = Table(kpi_data, colWidths=["25%"] * 4)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
    ]))
    elements += [kpi_table, Spacer(1, 6)]

    # Data quality breakdown
    elements.append(Paragraph("Data Quality Breakdown", h2_style))
    q_scores = quality.get("scores", {})
    q_rows = [["Dimension", "Score"]] + [[k.title(), f"{v}%"] for k, v in q_scores.items()]
    q_table = Table(q_rows, colWidths=["70%", "30%"])
    q_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), INK), ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_ROW]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements += [q_table, Spacer(1, 4)]

    # Daily summary (trailing 14 days to keep the PDF to a sane length)
    elements.append(Paragraph("Daily Summary (most recent 14 days)", h2_style))
    recent = summary.tail(14)
    sday_rows = [["Day", "Tx Count", "ETH Volume", "Avg Gas (gwei)", "Failed"]]
    for _, r in recent.iterrows():
        sday_rows.append([r["day"], str(int(r["tx_count"])), f"{round(r['total_eth_volume'], 3)}",
                           f"{round(r['avg_gas_price_gwei'], 1)}", str(int(r["failed_tx_count"]))])
    sday_table = Table(sday_rows, colWidths=["24%", "16%", "22%", "20%", "18%"], repeatRows=1)
    sday_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), INK), ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_ROW]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements += [sday_table, Spacer(1, 4)]

    # Anomalies
    anomalies = analytics.get("anomalies", [])[:10]
    elements.append(Paragraph(f"Anomalies Detected ({len(anomalies)} of {len(analytics.get('anomalies', []))} shown)", h2_style))
    if anomalies:
        a_rows = [["Hash", "Metric", "Value", "Z-score"]]
        for a in anomalies:
            a_rows.append([a["hash"][:14] + "...", a["metric"], str(a["value"]), str(a["z_score"])])
        a_table = Table(a_rows, colWidths=["40%", "20%", "20%", "20%"])
        a_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), INK), ("TEXTCOLOR", (0, 0), (-1, 0), GOLD),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_ROW]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TEXTCOLOR", (3, 1), (3, -1), colors.HexColor("#c4634a")),
        ]))
        elements.append(a_table)
    else:
        elements.append(Paragraph("No statistical anomalies detected in this run.", body_style))

    elements += [Spacer(1, 18),
                 Paragraph("Methodology: quality score = completeness(30%) + uniqueness(20%) + "
                           "validity(35%) + timeliness(15%). Anomalies flagged where |z-score| > 3 on "
                           "transaction value or gas price. Full methodology at /integrity-methodology.html",
                           muted_style),
                 Spacer(1, 10),
                 Paragraph("Made with love by Goldy — github.com/goldy2307", muted_style)]

    doc.build(elements)
    return buf.getvalue()